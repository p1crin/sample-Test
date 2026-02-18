#!/usr/bin/env python3
"""
ProofLink 総合テスト(IT2)試験書 生成スクリプト
性能テスト・負荷テスト・シナリオテストの試験項目書をExcelで生成する
"""

import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from datetime import datetime
import os

# === 共通スタイル定義 ===
HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
HEADER_FONT = Font(name="游ゴシック", size=10, bold=True, color="FFFFFF")
NORMAL_FONT = Font(name="游ゴシック", size=10)
BOLD_FONT = Font(name="游ゴシック", size=10, bold=True)
TITLE_FONT = Font(name="游ゴシック", size=14, bold=True)
THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)
WRAP_ALIGNMENT = Alignment(wrap_text=True, vertical="top")
CENTER_ALIGNMENT = Alignment(horizontal="center", vertical="center", wrap_text=True)
HEADER_ALIGNMENT = Alignment(horizontal="center", vertical="center", wrap_text=True)

# 列マッピング（既存フォーマットに合わせる）
# A:ID, E:テスト大項目, O:テスト中項目, Y:テスト小項目, AI:正常系/異常系
# AS:設計仕様, BE:テスト観点, BQ:前提条件, CD:テスト手順, CM:期待結果
# CZ:実行結果, DC:実施日, DF:実施者, DI:確認日, DL:確認者, DO:備考
COL_MAP = {
    "ID": "A",
    "テスト大項目": "E",
    "テスト中項目": "O",
    "テスト小項目": "Y",
    "正常系/異常系": "AI",
    "設計仕様": "AS",
    "テスト観点": "BE",
    "前提条件": "BQ",
    "テスト手順": "CD",
    "期待結果": "CM",
    "実行結果": "CZ",
    "実施日": "DC",
    "実施者": "DF",
    "確認日": "DI",
    "確認者": "DL",
    "備考": "DO",
}

# 列幅設定
COL_WIDTHS = {
    "A": 14, "B": 3, "C": 3, "D": 3,  # ID
    "E": 4, "F": 4, "G": 4, "H": 4, "I": 4, "J": 4, "K": 4, "L": 4, "M": 4, "N": 4,  # テスト大項目
    "O": 4, "P": 4, "Q": 4, "R": 4, "S": 4, "T": 4, "U": 4, "V": 4, "W": 4, "X": 4,  # テスト中項目
    "Y": 4, "Z": 4, "AA": 4, "AB": 4, "AC": 4, "AD": 4, "AE": 4, "AF": 4, "AG": 4, "AH": 4,  # テスト小項目
    "AI": 4, "AJ": 4, "AK": 4, "AL": 4,  # 正常系/異常系
    "AM": 4, "AN": 4, "AO": 4, "AP": 4, "AQ": 4, "AR": 4,  # (設計仕様の一部)
    "AS": 4, "AT": 4, "AU": 4, "AV": 4, "AW": 4, "AX": 4, "AY": 4, "AZ": 4,
    "BA": 4, "BB": 4, "BC": 4, "BD": 4,  # 設計仕様
    "BE": 4, "BF": 4, "BG": 4, "BH": 4, "BI": 4, "BJ": 4, "BK": 4, "BL": 4,
    "BM": 4, "BN": 4, "BO": 4, "BP": 4,  # テスト観点
    "BQ": 4, "BR": 4, "BS": 4, "BT": 4, "BU": 4, "BV": 4, "BW": 4, "BX": 4,
    "BY": 4, "BZ": 4, "CA": 4, "CB": 4, "CC": 4,  # 前提条件
    "CD": 4, "CE": 4, "CF": 4, "CG": 4, "CH": 4, "CI": 4, "CJ": 4, "CK": 4, "CL": 4,  # テスト手順
    "CM": 4, "CN": 4, "CO": 4, "CP": 4, "CQ": 4, "CR": 4, "CS": 4, "CT": 4,
    "CU": 4, "CV": 4, "CW": 4, "CX": 4, "CY": 4,  # 期待結果
    "CZ": 4, "DA": 4, "DB": 4,  # 実行結果
    "DC": 4, "DD": 4, "DE": 4,  # 実施日
    "DF": 4, "DG": 4, "DH": 4,  # 実施者
    "DI": 4, "DJ": 4, "DK": 4,  # 確認日
    "DL": 4, "DM": 4, "DN": 4,  # 確認者
    "DO": 4, "DP": 4, "DQ": 4, "DR": 4, "DS": 4, "DT": 4, "DU": 4, "DV": 4, "DW": 4,  # 備考
}

# ヘッダー行のセル結合範囲（行1: メタ情報, 行2: 値, 行4: カラムヘッダー）
HEADER_MERGES_ROW1 = [
    ("A1", "E1"),   # システム名
    ("F1", "S1"),   # ドキュメント名
    ("T1", "V1"),   # 画面ID
    ("W1", "AF1"),  # 対象機能名
    ("AG1", "AH1"), # Ver.
    ("AI1", "AL1"), # 作成日
    ("AM1", "AS1"), # 作成者
    ("AT1", "AW1"), # 最終更新日
    ("AX1", "BD1"), # 最終更新者
]

HEADER_MERGES_ROW2 = [
    ("A2", "E2"),   # ProofLink
    ("F2", "S2"),   # ドキュメント名値
    ("T2", "V2"),   # 画面ID値
    ("W2", "AF2"),  # 対象機能名値
    ("AG2", "AH2"), # Ver.値
    ("AI2", "AL2"), # 作成日値
    ("AM2", "AS2"), # 作成者値
    ("AT2", "AW2"), # 最終更新日値
    ("AX2", "BD2"), # 最終更新者値
]

COL_HEADER_MERGES = [
    ("A4", "D4"),   # ID
    ("E4", "N4"),   # テスト大項目
    ("O4", "X4"),   # テスト中項目
    ("Y4", "AH4"),  # テスト小項目
    ("AI4", "AR4"), # 正常系/異常系 (実際にはAI4:AL4だが、既存に合わせる)
    ("AS4", "BD4"), # 設計仕様
    ("BE4", "BP4"), # テスト観点
    ("BQ4", "CC4"), # 前提条件
    ("CD4", "CL4"), # テスト手順
    ("CM4", "CY4"), # 期待結果
    ("CZ4", "DB4"), # 実行結果
    ("DC4", "DE4"), # 実施日
    ("DF4", "DH4"), # 実施者
    ("DI4", "DK4"), # 確認日
    ("DL4", "DN4"), # 確認者
    ("DO4", "DW4"), # 備考
]


def col_to_num(col_str):
    """列文字を数値に変換 (A=1, B=2, ..., Z=26, AA=27, ...)"""
    result = 0
    for c in col_str:
        result = result * 26 + (ord(c) - ord('A') + 1)
    return result


def apply_border_to_range(ws, start_row, end_row, start_col, end_col):
    """指定範囲にボーダーを適用"""
    for row in range(start_row, end_row + 1):
        for col in range(start_col, end_col + 1):
            ws.cell(row=row, column=col).border = THIN_BORDER


def create_cover_sheet(wb, doc_name):
    """表紙シートを作成"""
    ws = wb.active
    ws.title = "表紙"

    # 列幅設定
    for col_letter in "ABCDEFGHIJKLMNOP":
        if col_letter in ("A", "C", "F", "G", "H"):
            ws.column_dimensions[col_letter].width = 8.8
        else:
            ws.column_dimensions[col_letter].width = 13.0

    # タイトル
    ws.merge_cells("G16:N18")
    cell = ws["G16"]
    cell.value = doc_name
    cell.font = TITLE_FONT
    cell.alignment = Alignment(horizontal="center", vertical="center")

    # メタ情報
    meta_rows = [
        ("G22", "H22", "Ver.", "I22", "N22", "1.0"),
        ("G23", "H23", "作成日", "I23", "N23", datetime(2026, 2, 19)),
        ("G24", "H24", "作成者", "I24", "N24", "Ｓｋｙ株式会社"),
        ("G25", "H25", "最終更新日", "I25", "N25", datetime(2026, 2, 19)),
        ("G26", "H26", "最終更新者", "I26", "N26", "Ｓｋｙ株式会社"),
        ("G27", "H27", "承認者", "I27", "N27", ""),
    ]
    for label_start, label_end, label, val_start, val_end, value in meta_rows:
        ws.merge_cells(f"{label_start}:{label_end}")
        ws[label_start].value = label
        ws[label_start].font = BOLD_FONT
        ws[label_start].border = THIN_BORDER
        ws[label_start.replace(label_start[0], chr(ord(label_start[0])+1), 1) if len(label_start) == 3 else label_end].border = THIN_BORDER
        ws.merge_cells(f"{val_start}:{val_end}")
        ws[val_start].value = value
        ws[val_start].font = NORMAL_FONT
        ws[val_start].border = THIN_BORDER
        # ボーダーを結合範囲に適用
        for r in range(int(label_start[1:]), int(label_start[1:]) + 1):
            for c_letter in ["G", "H", "I", "J", "K", "L", "M", "N"]:
                ws[f"{c_letter}{r}"].border = THIN_BORDER

    return ws


def create_revision_sheet(wb):
    """改版履歴シートを作成"""
    ws = wb.create_sheet("改版履歴")

    # ヘッダー
    headers = [("A1", "No"), ("B1", "Ver"), ("C1", "内容"), ("Q1", "作成者"), ("S1", "作成日")]
    ws.merge_cells("C1:P1")
    ws.merge_cells("Q1:R1")
    ws.merge_cells("S1:T1")

    for cell_ref, value in headers:
        ws[cell_ref].value = value
        ws[cell_ref].font = HEADER_FONT
        ws[cell_ref].fill = HEADER_FILL
        ws[cell_ref].border = THIN_BORDER
        ws[cell_ref].alignment = HEADER_ALIGNMENT

    # ヘッダー行のボーダー
    for col in range(1, 21):
        ws.cell(row=1, column=col).border = THIN_BORDER
        ws.cell(row=1, column=col).fill = HEADER_FILL
        ws.cell(row=1, column=col).font = HEADER_FONT

    # 初版
    ws["A2"].value = 1
    ws["B2"].value = "1.0"
    ws.merge_cells("C2:P2")
    ws["C2"].value = "初版作成"
    ws.merge_cells("Q2:R2")
    ws["Q2"].value = "Ｓｋｙ松石拓磨"
    ws.merge_cells("S2:T2")
    ws["S2"].value = datetime(2026, 2, 19)
    ws["S2"].number_format = "YYYY/M/D"
    for col in range(1, 21):
        ws.cell(row=2, column=col).border = THIN_BORDER
        ws.cell(row=2, column=col).font = NORMAL_FONT

    # 空行のテンプレート（3-15行）
    for row in range(3, 16):
        ws[f"A{row}"].value = f"=ROW()-1"
        ws.merge_cells(f"C{row}:P{row}")
        ws.merge_cells(f"Q{row}:R{row}")
        ws.merge_cells(f"S{row}:T{row}")
        for col in range(1, 21):
            ws.cell(row=row, column=col).border = THIN_BORDER

    ws.column_dimensions["A"].width = 8.8
    ws.column_dimensions["B"].width = 13.0
    for c in "CDEFGHIJKLMNOP":
        ws.column_dimensions[c].width = 13.0

    return ws


def setup_test_sheet(ws, screen_id, doc_name, target_name):
    """画面試験項目シートのヘッダーを設定"""
    # 列幅設定
    for col_letter, width in COL_WIDTHS.items():
        ws.column_dimensions[col_letter].width = width

    # 行1: メタ情報ラベル
    for start, end in HEADER_MERGES_ROW1:
        ws.merge_cells(f"{start}:{end}")

    labels_row1 = {
        "A1": "システム名", "F1": "ドキュメント名", "T1": "画面ID",
        "W1": "対象機能名", "AG1": "Ver.", "AI1": "作成日",
        "AM1": "作成者", "AT1": "最終更新日", "AX1": "最終更新者",
    }
    for cell_ref, value in labels_row1.items():
        cell = ws[cell_ref]
        cell.value = value
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGNMENT
        cell.border = THIN_BORDER

    # 行1のボーダーを全結合範囲に適用
    for start, end in HEADER_MERGES_ROW1:
        start_col = col_to_num(start.replace("1", ""))
        end_col = col_to_num(end.replace("1", ""))
        for c in range(start_col, end_col + 1):
            ws.cell(row=1, column=c).border = THIN_BORDER
            ws.cell(row=1, column=c).fill = HEADER_FILL
            ws.cell(row=1, column=c).font = HEADER_FONT

    # 行2: メタ情報値
    for start, end in HEADER_MERGES_ROW2:
        ws.merge_cells(f"{start}:{end}")

    values_row2 = {
        "A2": "ProofLink", "F2": doc_name, "T2": screen_id,
        "W2": target_name, "AG2": "1.0",
        "AI2": datetime(2026, 2, 19), "AM2": "Ｓｋｙ松石拓磨",
        "AT2": datetime(2026, 2, 19), "AX2": "Ｓｋｙ松石拓磨",
    }
    for cell_ref, value in values_row2.items():
        cell = ws[cell_ref]
        cell.value = value
        cell.font = NORMAL_FONT
        cell.alignment = WRAP_ALIGNMENT
        cell.border = THIN_BORDER
        if isinstance(value, datetime):
            cell.number_format = "M/D/YY"

    for start, end in HEADER_MERGES_ROW2:
        start_col = col_to_num(start.replace("2", ""))
        end_col = col_to_num(end.replace("2", ""))
        for c in range(start_col, end_col + 1):
            ws.cell(row=2, column=c).border = THIN_BORDER

    # 行3: 空行

    # 行4: カラムヘッダー
    col_headers = {
        "A4": "ID", "E4": "テスト大項目", "O4": "テスト中項目",
        "Y4": "テスト小項目", "AI4": "正常系/異常系", "AS4": "設計仕様",
        "BE4": "テスト観点", "BQ4": "前提条件", "CD4": "テスト手順",
        "CM4": "期待結果", "CZ4": "実行結果", "DC4": "実施日",
        "DF4": "実施者", "DI4": "確認日", "DL4": "確認者", "DO4": "備考",
    }
    for start, end in COL_HEADER_MERGES:
        ws.merge_cells(f"{start}:{end}")
        start_col = col_to_num(start.replace("4", ""))
        end_col = col_to_num(end.replace("4", ""))
        for c in range(start_col, end_col + 1):
            ws.cell(row=4, column=c).border = THIN_BORDER
            ws.cell(row=4, column=c).fill = HEADER_FILL
            ws.cell(row=4, column=c).font = HEADER_FONT

    for cell_ref, value in col_headers.items():
        cell = ws[cell_ref]
        cell.value = value
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGNMENT
        cell.border = THIN_BORDER


def write_test_row(ws, row, test_id, major, medium, minor, normal_abnormal,
                   spec, viewpoint, precondition, procedure, expected, note=""):
    """テスト項目の1行を書き込む（結合なし版、単一行）"""
    cell_data = {
        "A": test_id, "E": major, "O": medium, "Y": minor,
        "AI": normal_abnormal, "AS": spec, "BE": viewpoint,
        "BQ": precondition, "CD": procedure, "CM": expected, "DO": note,
    }
    for col_letter, value in cell_data.items():
        if value:
            cell = ws[f"{col_letter}{row}"]
            cell.value = value
            cell.font = NORMAL_FONT
            cell.alignment = WRAP_ALIGNMENT

    # ボーダーを全セルに適用
    last_col = col_to_num("DW")
    for c in range(1, last_col + 1):
        ws.cell(row=row, column=c).border = THIN_BORDER


def merge_cells_for_row(ws, row, start_row, col_start, col_end):
    """指定行範囲でセル結合する"""
    if start_row < row:
        start_col_letter = col_start
        end_col_letter = col_end
        ws.merge_cells(f"{start_col_letter}{start_row}:{end_col_letter}{row - 1}")


def write_test_items(ws, items, screen_id, test_type="IT2"):
    """テスト項目をシートに書き込む（セル結合対応）"""
    # 各セル結合用の列範囲定義
    merge_col_ranges = {
        "ID": ("A", "D"),
        "テスト大項目": ("E", "N"),
        "テスト中項目": ("O", "X"),
        "テスト小項目": ("Y", "AH"),
        "正常系/異常系": ("AI", "AR"),
        "設計仕様": ("AS", "BD"),
        "テスト観点": ("BE", "BP"),
        "前提条件": ("BQ", "CC"),
        "テスト手順": ("CD", "CL"),
        "期待結果": ("CM", "CY"),
        "実行結果": ("CZ", "DB"),
        "実施日": ("DC", "DE"),
        "実施者": ("DF", "DH"),
        "確認日": ("DI", "DK"),
        "確認者": ("DL", "DN"),
        "備考": ("DO", "DW"),
    }

    row = 5  # データ開始行
    test_num = 1
    prev_major = None
    prev_medium = None
    major_start_row = row
    medium_start_row = row

    for item in items:
        test_id = f"{screen_id}-{test_type}-{test_num}"
        major = item.get("major", "")
        medium = item.get("medium", "")
        minor = item.get("minor", "")
        normal_abnormal = item.get("type", "正常系")
        spec = item.get("spec", "")
        viewpoint = item.get("viewpoint", "")
        precondition = item.get("precondition", "")
        steps = item.get("steps", [])
        expected_results = item.get("expected", [])
        note = item.get("note", "")

        # 複数ステップがある場合、複数行にまたがる
        num_rows = max(len(steps), len(expected_results), 1)
        start_row = row

        for i in range(num_rows):
            # ボーダーを全セルに適用
            last_col = col_to_num("DW")
            for c in range(1, last_col + 1):
                ws.cell(row=row, column=c).border = THIN_BORDER

            # ステップ
            if i < len(steps):
                ws[f"CD{row}"].value = steps[i]
                ws[f"CD{row}"].font = NORMAL_FONT
                ws[f"CD{row}"].alignment = WRAP_ALIGNMENT

            # 期待結果
            if i < len(expected_results):
                ws[f"CM{row}"].value = expected_results[i]
                ws[f"CM{row}"].font = NORMAL_FONT
                ws[f"CM{row}"].alignment = WRAP_ALIGNMENT

            row += 1

        # 最初の行にデータを書き込み
        ws[f"A{start_row}"].value = test_id
        ws[f"A{start_row}"].font = NORMAL_FONT
        ws[f"A{start_row}"].alignment = WRAP_ALIGNMENT

        if major != prev_major:
            ws[f"E{start_row}"].value = major
            ws[f"E{start_row}"].font = NORMAL_FONT
            ws[f"E{start_row}"].alignment = WRAP_ALIGNMENT

        ws[f"O{start_row}"].value = medium
        ws[f"O{start_row}"].font = NORMAL_FONT
        ws[f"O{start_row}"].alignment = WRAP_ALIGNMENT

        if minor:
            ws[f"Y{start_row}"].value = minor
            ws[f"Y{start_row}"].font = NORMAL_FONT
            ws[f"Y{start_row}"].alignment = WRAP_ALIGNMENT

        ws[f"AI{start_row}"].value = normal_abnormal
        ws[f"AI{start_row}"].font = NORMAL_FONT
        ws[f"AI{start_row}"].alignment = WRAP_ALIGNMENT

        if spec:
            ws[f"AS{start_row}"].value = spec
            ws[f"AS{start_row}"].font = NORMAL_FONT
            ws[f"AS{start_row}"].alignment = WRAP_ALIGNMENT

        ws[f"BE{start_row}"].value = viewpoint
        ws[f"BE{start_row}"].font = NORMAL_FONT
        ws[f"BE{start_row}"].alignment = WRAP_ALIGNMENT

        if precondition:
            ws[f"BQ{start_row}"].value = precondition
            ws[f"BQ{start_row}"].font = NORMAL_FONT
            ws[f"BQ{start_row}"].alignment = WRAP_ALIGNMENT

        if note:
            ws[f"DO{start_row}"].value = note
            ws[f"DO{start_row}"].font = NORMAL_FONT
            ws[f"DO{start_row}"].alignment = WRAP_ALIGNMENT

        # 複数行のセル結合
        if num_rows > 1:
            for field, (col_s, col_e) in merge_col_ranges.items():
                if field not in ("テスト手順", "期待結果"):
                    ws.merge_cells(f"{col_s}{start_row}:{col_e}{start_row + num_rows - 1}")
                else:
                    # テスト手順と期待結果は1行ずつ結合
                    for r in range(start_row, start_row + num_rows):
                        ws.merge_cells(f"{col_s}{r}:{col_e}{r}")
        else:
            # 単一行でも列内結合は必要
            for field, (col_s, col_e) in merge_col_ranges.items():
                ws.merge_cells(f"{col_s}{start_row}:{col_e}{start_row}")

        prev_major = major
        test_num += 1

    # テスト大項目のセル結合処理
    _merge_major_items(ws, items, screen_id, test_type)

    return row


def _merge_major_items(ws, items, screen_id, test_type):
    """テスト大項目が同じ連続行をセル結合する"""
    row = 5
    prev_major = None
    major_start = 5
    merge_ranges = []

    for item in items:
        major = item.get("major", "")
        steps = item.get("steps", [])
        expected_results = item.get("expected", [])
        num_rows = max(len(steps), len(expected_results), 1)

        if major != prev_major and prev_major is not None:
            if row - major_start > 0:
                merge_ranges.append((major_start, row - 1))
            major_start = row
        elif major == prev_major:
            pass  # 継続

        prev_major = major
        row += num_rows

    # 最後のグループ
    if row - major_start > 0:
        merge_ranges.append((major_start, row - 1))

    # テスト大項目列のセル結合
    for start, end in merge_ranges:
        if end > start:
            # 既に個別行で結合済みなので、テスト大項目列のみ再結合
            # Note: openpyxlでは既存の結合を解除してから再結合する必要がある
            try:
                for r in range(start, end + 1):
                    try:
                        ws.unmerge_cells(f"E{r}:N{r}")
                    except (KeyError, ValueError):
                        pass
                ws.merge_cells(f"E{start}:N{end}")
            except (KeyError, ValueError):
                pass


# === テスト項目データ定義 ===

def get_performance_test_items():
    """性能テスト項目"""
    return [
        # === テストグループ複製 ===
        {
            "major": "テストグループ複製",
            "medium": "レスポンスタイム",
            "minor": "小規模グループ",
            "type": "正常系",
            "spec": "・テストグループ複製API\n・POST /api/test-groups/[groupId]",
            "viewpoint": "テストケース50件以下のテストグループ複製が3秒以内に完了すること",
            "precondition": "・テストケース50件のテストグループが存在すること\n・テスト内容が各テストケースに3件ずつ存在すること\n・添付ファイルが各テストケースに1件ずつ存在すること",
            "steps": ["1.テストグループ複製APIを実行する"],
            "expected": ["・レスポンスタイムが3秒以内であること\n・複製後のテストグループが正常に表示されること"],
            "note": "計測ツール: JMeter\n計測回数: 5回の平均値",
        },
        {
            "major": "テストグループ複製",
            "medium": "レスポンスタイム",
            "minor": "中規模グループ",
            "type": "正常系",
            "spec": "・テストグループ複製API\n・POST /api/test-groups/[groupId]",
            "viewpoint": "テストケース200件のテストグループ複製が10秒以内に完了すること",
            "precondition": "・テストケース200件のテストグループが存在すること\n・テスト内容が各テストケースに5件ずつ存在すること\n・添付ファイルが各テストケースに2件ずつ存在すること",
            "steps": ["1.テストグループ複製APIを実行する"],
            "expected": ["・レスポンスタイムが10秒以内であること\n・複製後のデータ件数が元のグループと一致すること"],
            "note": "計測ツール: JMeter\n計測回数: 5回の平均値",
        },
        {
            "major": "テストグループ複製",
            "medium": "レスポンスタイム",
            "minor": "大規模グループ",
            "type": "正常系",
            "spec": "・テストグループ複製API\n・POST /api/test-groups/[groupId]",
            "viewpoint": "テストケース500件のテストグループ複製が30秒以内に完了すること",
            "precondition": "・テストケース500件のテストグループが存在すること\n・テスト内容が各テストケースに10件ずつ存在すること\n・添付ファイルが各テストケースに3件ずつ存在すること\n・エビデンスファイルが各テスト結果に2件ずつ存在すること",
            "steps": ["1.テストグループ複製APIを実行する"],
            "expected": ["・レスポンスタイムが30秒以内であること\n・複製後のデータ整合性が保たれていること（テストケース、テスト内容、ファイル、エビデンス全て）"],
            "note": "計測ツール: JMeter\n計測回数: 5回の平均値",
        },
        {
            "major": "テストグループ複製",
            "medium": "データ整合性",
            "minor": "",
            "type": "正常系",
            "spec": "・テストグループ複製API\n・POST /api/test-groups/[groupId]",
            "viewpoint": "大規模テストグループ複製後のデータ整合性が保たれていること",
            "precondition": "・テストケース500件のテストグループが存在すること\n・全テーブル（tt_test_cases, tt_test_contents, tt_test_case_files, tt_test_results, tt_test_results_history, tt_test_evidences, tt_test_group_tags）にデータが存在すること",
            "steps": ["1.テストグループ複製APIを実行する", "2.複製元と複製先のデータ件数を比較する", "3.複製先のS3ファイルパスが正しく設定されていることを確認する"],
            "expected": ["・複製が正常に完了すること", "・全テーブルのレコード数が複製元と一致すること", "・S3上のファイルパスが新グループIDのディレクトリに格納されていること"],
        },
        # === テストグループ集計 ===
        {
            "major": "テストグループ集計",
            "medium": "レスポンスタイム",
            "minor": "小規模グループ",
            "type": "正常系",
            "spec": "・テスト集計API\n・GET /api/test-groups/[groupId]/report-data",
            "viewpoint": "テストケース50件以下のテストグループ集計が1秒以内に完了すること",
            "precondition": "・テストケース50件のテストグループが存在すること\n・テスト結果が入力済みであること",
            "steps": ["1.テスト集計APIを実行する"],
            "expected": ["・レスポンスタイムが1秒以内であること\n・集計結果（total_items, completed_items, ok_items, ng_items等）が正しいこと"],
            "note": "計測ツール: JMeter\n計測回数: 5回の平均値",
        },
        {
            "major": "テストグループ集計",
            "medium": "レスポンスタイム",
            "minor": "中規模グループ",
            "type": "正常系",
            "spec": "・テスト集計API\n・GET /api/test-groups/[groupId]/report-data",
            "viewpoint": "テストケース200件のテストグループ集計が3秒以内に完了すること",
            "precondition": "・テストケース200件のテストグループが存在すること\n・テスト内容が各テストケースに5件ずつ存在すること\n・テスト結果が入力済みであること",
            "steps": ["1.テスト集計APIを実行する"],
            "expected": ["・レスポンスタイムが3秒以内であること\n・first_layer, second_layer別の集計結果が正しいこと"],
            "note": "計測ツール: JMeter\n計測回数: 5回の平均値",
        },
        {
            "major": "テストグループ集計",
            "medium": "レスポンスタイム",
            "minor": "大規模グループ",
            "type": "正常系",
            "spec": "・テスト集計API\n・GET /api/test-groups/[groupId]/report-data",
            "viewpoint": "テストケース500件のテストグループ集計が5秒以内に完了すること",
            "precondition": "・テストケース500件のテストグループが存在すること\n・テスト内容が各テストケースに10件ずつ存在すること\n・テスト結果が全件入力済みであること",
            "steps": ["1.テスト集計APIを実行する"],
            "expected": ["・レスポンスタイムが5秒以内であること\n・ok_rate, progress_rateの計算結果が手動計算値と一致すること"],
            "note": "計測ツール: JMeter\n計測回数: 5回の平均値",
        },
        {
            "major": "テストグループ集計",
            "medium": "日次レポート",
            "minor": "",
            "type": "正常系",
            "spec": "・日次レポートAPI\n・GET /api/test-groups/[groupId]/daily-report-data",
            "viewpoint": "日次レポートデータの取得が3秒以内に完了すること",
            "precondition": "・テストケース500件のテストグループが存在すること\n・過去30日間のテスト結果履歴が存在すること",
            "steps": ["1.日次レポートAPIを実行する"],
            "expected": ["・レスポンスタイムが3秒以内であること\n・日付別の集計データが正しいこと"],
            "note": "計測ツール: JMeter",
        },
        # === テストインポートバッチ ===
        {
            "major": "テストインポートバッチ",
            "medium": "処理時間",
            "minor": "小規模インポート",
            "type": "正常系",
            "spec": "・テストケースインポートバッチ\n・batch/src/test-case-import.ts",
            "viewpoint": "テストケース50件のCSVインポートが30秒以内に完了すること",
            "precondition": "・テストケース50件分のCSVファイルを含むZIPファイルが用意されていること\n・添付ファイルが10件含まれていること\n・インポート先のテストグループが存在すること",
            "steps": ["1.テストケースインポートバッチを実行する"],
            "expected": ["・処理が30秒以内に完了すること\n・全50件のテストケースがDBに正しく登録されていること\n・添付ファイルがS3に正しくアップロードされていること"],
            "note": "AWS Batch環境で実行",
        },
        {
            "major": "テストインポートバッチ",
            "medium": "処理時間",
            "minor": "中規模インポート",
            "type": "正常系",
            "spec": "・テストケースインポートバッチ\n・batch/src/test-case-import.ts",
            "viewpoint": "テストケース200件のCSVインポートが2分以内に完了すること",
            "precondition": "・テストケース200件分のCSVファイルを含むZIPファイルが用意されていること\n・添付ファイルが50件含まれていること\n・インポート先のテストグループが存在すること",
            "steps": ["1.テストケースインポートバッチを実行する"],
            "expected": ["・処理が2分以内に完了すること\n・全200件のテストケースがDBに正しく登録されていること"],
            "note": "AWS Batch環境で実行",
        },
        {
            "major": "テストインポートバッチ",
            "medium": "処理時間",
            "minor": "大規模インポート",
            "type": "正常系",
            "spec": "・テストケースインポートバッチ\n・batch/src/test-case-import.ts",
            "viewpoint": "テストケース500件のCSVインポートが5分以内に完了すること",
            "precondition": "・テストケース500件分のCSVファイルを含むZIPファイルが用意されていること\n・添付ファイルが100件含まれていること\n・インポート先のテストグループが存在すること",
            "steps": ["1.テストケースインポートバッチを実行する"],
            "expected": ["・処理が5分以内に完了すること\n・全500件のテストケースがDBに正しく登録されていること\n・結果ファイル（JSON/CSV）がS3に出力されていること"],
            "note": "AWS Batch環境で実行",
        },
        {
            "major": "テストインポートバッチ",
            "medium": "トランザクション整合性",
            "minor": "途中エラー時ロールバック",
            "type": "異常系",
            "spec": "・テストケースインポートバッチ\n・batch/src/test-case-import.ts",
            "viewpoint": "インポート途中でエラーが発生した場合にトランザクションがロールバックされること",
            "precondition": "・テストケース100件分のCSVファイルを含むZIPファイルが用意されていること\n・50件目のレコードに不正データ（重複TID等）が含まれていること",
            "steps": ["1.不正データを含むZIPファイルでインポートバッチを実行する"],
            "expected": ["・エラーが検出されインポートが中断すること\n・DBにレコードが1件も追加されていないこと（全件ロールバック）\n・エラー結果ファイルにエラー内容が記録されていること"],
        },
        # === ユーザインポートバッチ ===
        {
            "major": "ユーザインポートバッチ",
            "medium": "処理時間",
            "minor": "大規模インポート",
            "type": "正常系",
            "spec": "・ユーザインポートバッチ\n・batch/src/user-import.ts",
            "viewpoint": "ユーザ100件のCSVインポートが1分以内に完了すること",
            "precondition": "・ユーザ100件分のCSVファイルがS3にアップロードされていること\n・各ユーザにタグが2-3件設定されていること",
            "steps": ["1.ユーザインポートバッチを実行する"],
            "expected": ["・処理が1分以内に完了すること\n・全100件のユーザがDBに正しく登録されていること\n・パスワードがbcryptでハッシュ化されていること\n・タグが正しく紐付けられていること"],
            "note": "AWS Batch環境で実行",
        },
        # === テストグループ一覧 ===
        {
            "major": "テストグループ一覧表示",
            "medium": "レスポンスタイム",
            "minor": "",
            "type": "正常系",
            "spec": "・テストグループ一覧API\n・GET /api/test-groups",
            "viewpoint": "テストグループ100件の一覧表示が2秒以内に完了すること",
            "precondition": "・テストグループが100件登録されていること\n・各グループにテストケースが存在すること",
            "steps": ["1.テストグループ一覧APIを実行する"],
            "expected": ["・レスポンスタイムが2秒以内であること\n・全100件のテストグループが正しく表示されること"],
            "note": "計測ツール: JMeter",
        },
        # === テストケース一覧 ===
        {
            "major": "テストケース一覧表示",
            "medium": "レスポンスタイム",
            "minor": "",
            "type": "正常系",
            "spec": "・テストケース一覧API\n・GET /api/test-groups/[groupId]/cases",
            "viewpoint": "テストケース500件の一覧表示が3秒以内に完了すること",
            "precondition": "・テストケース500件のテストグループが存在すること",
            "steps": ["1.テストケース一覧APIを実行する"],
            "expected": ["・レスポンスタイムが3秒以内であること\n・全500件のテストケースが正しく表示されること"],
            "note": "計測ツール: JMeter",
        },
        # === 認証 ===
        {
            "major": "認証処理",
            "medium": "レスポンスタイム",
            "minor": "",
            "type": "正常系",
            "spec": "・認証API\n・POST /api/auth/[...nextauth]",
            "viewpoint": "ログイン処理が2秒以内に完了すること",
            "precondition": "・有効なユーザアカウントが存在すること",
            "steps": ["1.ログインAPIを実行する"],
            "expected": ["・レスポンスタイムが2秒以内であること\n・JWTトークンが正しく発行されること"],
            "note": "計測ツール: JMeter",
        },
        # === ファイルアップロード ===
        {
            "major": "ファイルアップロード",
            "medium": "レスポンスタイム",
            "minor": "エビデンスファイル",
            "type": "正常系",
            "spec": "・ファイルアップロードAPI\n・POST /api/files/evidences",
            "viewpoint": "10MBのエビデンスファイルアップロードが5秒以内に完了すること",
            "precondition": "・10MBの画像ファイルが用意されていること\n・S3バケットが正しく設定されていること",
            "steps": ["1.エビデンスファイルアップロードAPIを実行する"],
            "expected": ["・レスポンスタイムが5秒以内であること\n・ファイルがS3に正しく保存されること"],
            "note": "計測ツール: JMeter",
        },
    ]


def get_load_test_items():
    """負荷テスト項目"""
    return [
        # === 同時接続テスト ===
        {
            "major": "同時接続テスト",
            "medium": "テストグループ一覧",
            "minor": "10ユーザ同時アクセス",
            "type": "正常系",
            "spec": "・テストグループ一覧API\n・GET /api/test-groups",
            "viewpoint": "10ユーザが同時にテストグループ一覧にアクセスした場合、全リクエストが3秒以内に完了すること",
            "precondition": "・10ユーザ分のアカウントが存在すること\n・テストグループが100件登録されていること\n・JMeterで10スレッドを同時実行する設定であること",
            "steps": ["1.JMeterで10スレッドを同時起動してテストグループ一覧APIにアクセスする"],
            "expected": ["・全リクエストの95パーセンタイルレスポンスタイムが3秒以内であること\n・エラーレートが0%であること"],
            "note": "JMeter Thread Group: 10 threads, Ramp-up: 1s",
        },
        {
            "major": "同時接続テスト",
            "medium": "テストグループ一覧",
            "minor": "30ユーザ同時アクセス",
            "type": "正常系",
            "spec": "・テストグループ一覧API\n・GET /api/test-groups",
            "viewpoint": "30ユーザが同時にテストグループ一覧にアクセスした場合、全リクエストが5秒以内に完了すること",
            "precondition": "・30ユーザ分のアカウントが存在すること\n・テストグループが100件登録されていること\n・JMeterで30スレッドを同時実行する設定であること",
            "steps": ["1.JMeterで30スレッドを同時起動してテストグループ一覧APIにアクセスする"],
            "expected": ["・全リクエストの95パーセンタイルレスポンスタイムが5秒以内であること\n・エラーレートが1%未満であること"],
            "note": "JMeter Thread Group: 30 threads, Ramp-up: 3s",
        },
        {
            "major": "同時接続テスト",
            "medium": "テストグループ一覧",
            "minor": "50ユーザ同時アクセス",
            "type": "正常系",
            "spec": "・テストグループ一覧API\n・GET /api/test-groups",
            "viewpoint": "50ユーザが同時にテストグループ一覧にアクセスした場合のレスポンスタイムとエラーレートを確認する",
            "precondition": "・50ユーザ分のアカウントが存在すること\n・テストグループが100件登録されていること\n・JMeterで50スレッドを同時実行する設定であること",
            "steps": ["1.JMeterで50スレッドを同時起動してテストグループ一覧APIにアクセスする"],
            "expected": ["・全リクエストの95パーセンタイルレスポンスタイムが10秒以内であること\n・エラーレートが5%未満であること"],
            "note": "JMeter Thread Group: 50 threads, Ramp-up: 5s",
        },
        # === テストグループ複製の同時実行 ===
        {
            "major": "同時接続テスト",
            "medium": "テストグループ複製",
            "minor": "3ユーザ同時複製",
            "type": "正常系",
            "spec": "・テストグループ複製API\n・POST /api/test-groups/[groupId]",
            "viewpoint": "3ユーザが同時にテストグループ複製を実行した場合、全処理が正常に完了すること",
            "precondition": "・テストケース100件のテストグループが3つ存在すること\n・3ユーザ分のアカウントが存在すること\n・JMeterで3スレッドを同時実行する設定であること",
            "steps": ["1.JMeterで3スレッドを同時起動して異なるテストグループの複製APIを実行する"],
            "expected": ["・全リクエストが正常に完了すること（HTTPステータス200）\n・各複製先のテストグループのデータ整合性が保たれていること\n・デッドロックが発生しないこと"],
            "note": "トランザクション競合に注意",
        },
        {
            "major": "同時接続テスト",
            "medium": "テストグループ複製",
            "minor": "同一グループ同時複製",
            "type": "異常系",
            "spec": "・テストグループ複製API\n・POST /api/test-groups/[groupId]",
            "viewpoint": "同一テストグループに対して3ユーザが同時に複製を実行した場合の排他制御が正しく動作すること",
            "precondition": "・テストケース100件のテストグループが1つ存在すること\n・3ユーザ分のアカウントが存在すること\n・JMeterで3スレッドが同一グループIDに対して複製を実行する設定であること",
            "steps": ["1.JMeterで3スレッドを同時起動して同一テストグループの複製APIを実行する"],
            "expected": ["・全リクエストが完了すること（成功またはエラー）\n・データ不整合が発生しないこと\n・複製されたグループのデータが正しいこと"],
            "note": "排他制御・デッドロック確認",
        },
        # === テスト集計の同時アクセス ===
        {
            "major": "同時接続テスト",
            "medium": "テストグループ集計",
            "minor": "10ユーザ同時集計",
            "type": "正常系",
            "spec": "・テスト集計API\n・GET /api/test-groups/[groupId]/report-data",
            "viewpoint": "10ユーザが同時に集計APIにアクセスした場合、全リクエストが5秒以内に完了すること",
            "precondition": "・テストケース500件のテストグループが存在すること\n・10ユーザ分のアカウントが存在すること",
            "steps": ["1.JMeterで10スレッドを同時起動して集計APIにアクセスする"],
            "expected": ["・全リクエストの95パーセンタイルレスポンスタイムが5秒以内であること\n・全リクエストの集計結果が同一であること"],
            "note": "JMeter Thread Group: 10 threads",
        },
        # === 持続負荷テスト ===
        {
            "major": "持続負荷テスト",
            "medium": "テストグループ一覧",
            "minor": "30分間持続負荷",
            "type": "正常系",
            "spec": "・テストグループ一覧API\n・GET /api/test-groups",
            "viewpoint": "30分間継続して負荷をかけた場合にレスポンスタイムが劣化しないこと",
            "precondition": "・10ユーザ分のアカウントが存在すること\n・テストグループが100件登録されていること\n・JMeterで10スレッド×30分間のループ設定であること",
            "steps": ["1.JMeterで10スレッドを30分間継続実行する"],
            "expected": ["・30分間を通じて95パーセンタイルレスポンスタイムが5秒以内を維持すること\n・エラーレートが1%未満であること\n・メモリリークの兆候がないこと（CloudWatchで確認）"],
            "note": "JMeter Duration: 1800s\nCloudWatchでメモリ・CPU使用率を監視",
        },
        {
            "major": "持続負荷テスト",
            "medium": "混合シナリオ",
            "minor": "60分間持続負荷",
            "type": "正常系",
            "spec": "・複数API混合\n・テストグループ一覧、テストケース一覧、集計、ファイルアップロード",
            "viewpoint": "60分間の混合負荷テストでシステムが安定動作すること",
            "precondition": "・20ユーザ分のアカウントが存在すること\n・十分なテストデータが登録されていること\n・JMeterで混合シナリオ（一覧40%, 詳細30%, 集計20%, ファイル10%）の設定であること",
            "steps": ["1.JMeterで20スレッドの混合シナリオを60分間継続実行する"],
            "expected": ["・60分間を通じてシステムが安定動作すること\n・95パーセンタイルレスポンスタイムが各API基準値以内であること\n・ECSタスクの再起動が発生しないこと\n・RDSのCPU使用率が80%を超えないこと"],
            "note": "JMeter Duration: 3600s\nCloudWatch, RDS Performance Insightsで監視",
        },
        # === スパイクテスト ===
        {
            "major": "スパイクテスト",
            "medium": "急激な負荷増加",
            "minor": "",
            "type": "正常系",
            "spec": "・テストグループ一覧API\n・GET /api/test-groups",
            "viewpoint": "急激な負荷増加時にシステムがダウンせず、負荷軽減後に正常に復帰すること",
            "precondition": "・50ユーザ分のアカウントが存在すること\n・テストグループが100件登録されていること\n・JMeterで段階的負荷（5→50→5ユーザ）の設定であること",
            "steps": ["1.JMeterで5スレッドから開始し、1分後に50スレッドに急増させ、さらに1分後に5スレッドに戻す"],
            "expected": ["・急激な負荷増加時にHTTP 5xxエラーが発生しないこと\n・負荷軽減後にレスポンスタイムが通常レベルに復帰すること\n・ALBのヘルスチェックが失敗しないこと"],
            "note": "JMeter Ultimate Thread Group使用",
        },
        # === バッチ処理中の負荷 ===
        {
            "major": "バッチ処理中の負荷テスト",
            "medium": "テストインポートバッチ実行中",
            "minor": "Web操作並行",
            "type": "正常系",
            "spec": "・テストケースインポートバッチ\n・テストグループ一覧API",
            "viewpoint": "テストインポートバッチ実行中に他ユーザのWeb操作が影響を受けないこと",
            "precondition": "・テストケース500件のインポートバッチが実行中であること\n・10ユーザ分のアカウントが存在すること\n・JMeterで10スレッドの一覧・詳細アクセスシナリオが設定されていること",
            "steps": ["1.テストインポートバッチを実行開始する", "2.バッチ実行中にJMeterで10スレッドのWeb操作シナリオを実行する"],
            "expected": ["・バッチ処理が実行中であること", "・Web操作のレスポンスタイムがバッチ非実行時と比較して2倍以内であること\n・Web操作でエラーが発生しないこと"],
            "note": "AWS Batchは別コンテナで実行されるため影響は限定的だが確認が必要",
        },
        # === DB接続プール ===
        {
            "major": "DB接続プールテスト",
            "medium": "接続枯渇",
            "minor": "",
            "type": "異常系",
            "spec": "・PostgreSQL接続プール\n・Prismaクライアント設定",
            "viewpoint": "大量同時リクエスト時にDB接続プールが枯渇しないこと、または適切にエラーハンドリングされること",
            "precondition": "・50ユーザ分のアカウントが存在すること\n・JMeterで50スレッドの高頻度リクエスト設定であること\n・Prismaの接続プール上限を確認済みであること",
            "steps": ["1.JMeterで50スレッドを同時起動し、高頻度でDB参照APIにアクセスする"],
            "expected": ["・接続プールが枯渇した場合、適切なエラーメッセージが返されること\n・システム全体がハングアップしないこと\n・負荷軽減後にDB接続が正常に回復すること"],
            "note": "CloudWatch RDS接続数を監視",
        },
    ]


def get_scenario_test_items():
    """シナリオテスト項目"""
    return [
        # === E2Eシナリオ: テストグループ管理 ===
        {
            "major": "テストグループ管理シナリオ",
            "medium": "テストグループの作成から集計まで",
            "minor": "",
            "type": "正常系",
            "spec": "・テストグループ管理全般\n・テストグループCRUD API\n・テスト集計API",
            "viewpoint": "テストグループの作成→テストケース追加→テスト実施→集計という一連の業務フローが正常に完了すること",
            "precondition": "・システム管理者アカウントでログイン済みであること\n・テスト用のタグが登録されていること",
            "steps": [
                "1.テストグループ一覧画面から新規テストグループを作成する（OEM、車種、イベント、仕向地等を入力）",
                "2.作成したテストグループのテストケース一覧画面に遷移する",
                "3.テストケースを5件手動で追加する（TID、第1層〜第4層、目的、確認観点、テスト手順を入力）",
                "4.各テストケースにテスト内容（テストケース、期待値）を3件ずつ追加する",
                "5.各テスト内容のテスト結果を入力する（結果、判定、実施日、ソフトVer等）",
                "6.テスト集計画面で集計結果を確認する",
            ],
            "expected": [
                "・テストグループが正常に作成され、テストグループ一覧に表示されること",
                "・テストケース一覧画面が正常に表示されること",
                "・テストケースが正常に登録され一覧に表示されること",
                "・テスト内容が正常に追加され表示されること",
                "・テスト結果が正常に保存されること",
                "・集計結果が入力したテスト結果と一致すること（OK数、NG数、進捗率等）",
            ],
        },
        # === E2Eシナリオ: テストグループ複製と編集 ===
        {
            "major": "テストグループ管理シナリオ",
            "medium": "テストグループ複製と差分編集",
            "minor": "",
            "type": "正常系",
            "spec": "・テストグループ複製API\n・テストケース編集API",
            "viewpoint": "既存テストグループを複製し、複製先のテストケースを編集する業務フローが正常に完了すること",
            "precondition": "・テストケース10件のテストグループが存在すること\n・テスト管理者アカウントでログイン済みであること",
            "steps": [
                "1.テストグループ一覧から対象グループの複製画面に遷移する",
                "2.複製先の情報を入力して複製を実行する",
                "3.複製されたテストグループのテストケース一覧を開く",
                "4.複製されたテストケースの第1層を編集して保存する",
                "5.複製元のテストグループのテストケースが変更されていないことを確認する",
            ],
            "expected": [
                "・テストグループ複製画面が正常に表示されること",
                "・複製が正常に完了し、新しいテストグループが作成されること",
                "・複製されたテストケースが元のグループと同じ内容で表示されること",
                "・編集した内容が正しく保存されること",
                "・複製元のテストケースが変更されていないこと（データ独立性の確認）",
            ],
        },
        # === E2Eシナリオ: テストケースインポート ===
        {
            "major": "テストインポートシナリオ",
            "medium": "CSVインポートからテスト実施まで",
            "minor": "",
            "type": "正常系",
            "spec": "・テストケースインポートバッチ\n・テストケース編集API\n・テスト結果入力API",
            "viewpoint": "CSVファイルによるテストケース一括インポートからテスト実施・結果入力までの一連のフローが正常に完了すること",
            "precondition": "・テストグループが作成済みであること\n・テストケース20件分のCSVと添付ファイルを含むZIPが準備されていること\n・テスト管理者アカウントでログイン済みであること",
            "steps": [
                "1.テストケースインポート画面からZIPファイルをアップロードする",
                "2.インポート結果一覧画面でインポートの完了を確認する",
                "3.テストケース一覧画面でインポートされたテストケースを確認する",
                "4.インポートされたテストケースの詳細を開き、添付ファイル（制御仕様書、データフロー）が正しくアップロードされていることを確認する",
                "5.テスト結果入力画面から結果を入力して保存する",
                "6.テスト集計画面で集計結果を確認する",
            ],
            "expected": [
                "・ファイルアップロードが正常に完了し、バッチジョブが開始されること",
                "・インポート結果が「成功」と表示され、件数が20件であること",
                "・20件のテストケースが一覧に表示されること",
                "・制御仕様書とデータフローのファイルが正しく表示されること",
                "・テスト結果が正常に保存され、結果履歴に記録されること",
                "・集計結果が入力した結果と一致すること",
            ],
        },
        # === E2Eシナリオ: インポートエラーハンドリング ===
        {
            "major": "テストインポートシナリオ",
            "medium": "不正データインポート時のエラーハンドリング",
            "minor": "",
            "type": "異常系",
            "spec": "・テストケースインポートバッチ\n・インポート結果API",
            "viewpoint": "不正なCSVデータを含むZIPファイルをインポートした場合、適切なエラーが表示されること",
            "precondition": "・テストグループが作成済みであること\n・不正データ（重複TID、必須項目欠落等）を含むCSVのZIPが準備されていること\n・テスト管理者アカウントでログイン済みであること",
            "steps": [
                "1.テストケースインポート画面から不正データを含むZIPファイルをアップロードする",
                "2.インポート結果一覧画面でインポートの完了を確認する",
                "3.インポート結果の詳細画面でエラー内容を確認する",
                "4.テストケース一覧画面で不正データが登録されていないことを確認する",
            ],
            "expected": [
                "・ファイルアップロードが正常に完了し、バッチジョブが開始されること",
                "・インポート結果が「エラー」と表示されること",
                "・エラー詳細にエラー原因（重複TID、必須項目欠落等）が表示されること",
                "・テストケースが1件も追加されていないこと（ロールバック確認）",
            ],
        },
        # === E2Eシナリオ: ユーザ管理 ===
        {
            "major": "ユーザ管理シナリオ",
            "medium": "ユーザCSVインポートとログイン",
            "minor": "",
            "type": "正常系",
            "spec": "・ユーザインポートバッチ\n・認証API\n・ユーザ管理API",
            "viewpoint": "CSVファイルによるユーザ一括インポート後、インポートされたユーザがログインして操作できること",
            "precondition": "・システム管理者アカウントでログイン済みであること\n・ユーザ5件分のCSVファイルが準備されていること\n・CSVに各ユーザのロール（管理者、テスト管理者、一般）が設定されていること",
            "steps": [
                "1.ユーザインポート実行画面からCSVファイルをアップロードする",
                "2.インポート結果一覧画面でインポートの完了を確認する",
                "3.ユーザ一覧画面でインポートされたユーザが表示されることを確認する",
                "4.インポートされたユーザでログインする",
                "5.ログイン後のサイドバーメニューがロールに応じて正しく表示されることを確認する",
            ],
            "expected": [
                "・ファイルアップロードが正常に完了し、バッチジョブが開始されること",
                "・インポート結果が「成功」と表示され、件数が5件であること",
                "・5件のユーザが一覧に表示されること",
                "・CSVに設定されたパスワードでログインできること",
                "・管理者はシステム管理者用メニュー含む全メニューが表示されること\n・テスト管理者はインポート管理含むメニューが表示されること\n・一般ユーザはテスト管理メニューのみ表示されること",
            ],
        },
        # === E2Eシナリオ: 権限制御 ===
        {
            "major": "権限制御シナリオ",
            "medium": "ロール別アクセス制御",
            "minor": "",
            "type": "正常系",
            "spec": "・認証・認可全般\n・ユーザロール制御",
            "viewpoint": "各ロール（管理者、テスト管理者、一般）のアクセス制御が正しく動作すること",
            "precondition": "・各ロールのユーザアカウントが存在すること\n・テストグループにテスト設計者、テスト実施者、テスト閲覧者タグが設定されていること",
            "steps": [
                "1.一般ユーザでログインし、ユーザ管理画面（/user）にアクセスする",
                "2.テスト管理者でログインし、ユーザ管理画面（/user）にアクセスする",
                "3.テスト閲覧者タグのユーザでログインし、テストケース編集画面にアクセスする",
                "4.テスト実施者タグのユーザでログインし、テスト結果入力画面にアクセスする",
                "5.システム管理者でログインし、全画面にアクセスする",
            ],
            "expected": [
                "・アクセス権限エラー画面が表示されること",
                "・アクセス権限エラー画面が表示されること",
                "・アクセス権限エラー画面が表示されること（テスト設計者タグ未保持のため）",
                "・テスト結果入力画面が正常に表示されること",
                "・全画面が正常に表示され、操作可能であること",
            ],
        },
        # === E2Eシナリオ: テスト結果の履歴管理 ===
        {
            "major": "テスト結果管理シナリオ",
            "medium": "テスト結果の複数回入力と履歴",
            "minor": "",
            "type": "正常系",
            "spec": "・テスト結果入力API\n・テスト結果履歴API",
            "viewpoint": "同一テスト内容に対してテスト結果を複数回入力した場合、履歴が正しく管理されること",
            "precondition": "・テストケースとテスト内容が登録済みであること\n・テスト実施者タグのユーザでログイン済みであること",
            "steps": [
                "1.テスト結果入力画面でテスト結果を入力して保存する（1回目: OK）",
                "2.同じテスト内容のテスト結果を再度入力して保存する（2回目: NG）",
                "3.テスト結果詳細画面で結果履歴を確認する",
                "4.集計画面で最新の結果が反映されていることを確認する",
            ],
            "expected": [
                "・テスト結果（OK）が正常に保存されること",
                "・テスト結果（NG）が正常に保存されること",
                "・結果履歴に2件の記録が表示され、history_countが正しいこと\n・最新の結果がNGであること",
                "・集計結果にNGとしてカウントされていること",
            ],
        },
        # === E2Eシナリオ: パスワード変更 ===
        {
            "major": "ユーザ管理シナリオ",
            "medium": "パスワード変更フロー",
            "minor": "",
            "type": "正常系",
            "spec": "・パスワード変更API\n・POST /api/auth/change-password",
            "viewpoint": "パスワード変更後に新しいパスワードでログインできること",
            "precondition": "・一般ユーザアカウントでログイン済みであること",
            "steps": [
                "1.パスワード変更画面で現在のパスワードと新しいパスワードを入力して変更する",
                "2.ログアウトする",
                "3.新しいパスワードでログインする",
                "4.旧パスワードでログインを試みる",
            ],
            "expected": [
                "・パスワード変更が正常に完了すること",
                "・ログアウトが正常に完了すること",
                "・新しいパスワードでログインが成功すること",
                "・旧パスワードでログインが失敗し、エラーメッセージが表示されること",
            ],
        },
        # === E2Eシナリオ: エビデンスファイル管理 ===
        {
            "major": "テスト結果管理シナリオ",
            "medium": "エビデンスファイルのアップロードと参照",
            "minor": "",
            "type": "正常系",
            "spec": "・ファイルアップロードAPI\n・ファイル参照API",
            "viewpoint": "テスト結果にエビデンスファイルをアップロードし、後から参照できること",
            "precondition": "・テストケースとテスト内容が登録済みであること\n・テスト実施者タグのユーザでログイン済みであること\n・画像ファイル（PNG, JPG）とPDFファイルが準備されていること",
            "steps": [
                "1.テスト結果入力画面でエビデンスとして画像ファイルをアップロードする",
                "2.テスト結果入力画面でエビデンスとしてPDFファイルをアップロードする",
                "3.テスト結果詳細画面でアップロードしたエビデンスを確認する",
                "4.エビデンスファイルをダウンロードして内容を確認する",
            ],
            "expected": [
                "・画像ファイルが正常にアップロードされ、サムネイルが表示されること",
                "・PDFファイルが正常にアップロードされ、ファイル名が表示されること",
                "・アップロードした全ファイルが一覧表示されること",
                "・ダウンロードしたファイルが元のファイルと同一であること（S3プリサインドURL経由）",
            ],
        },
        # === E2Eシナリオ: テストグループ削除 ===
        {
            "major": "テストグループ管理シナリオ",
            "medium": "テストグループ削除時のカスケード",
            "minor": "",
            "type": "正常系",
            "spec": "・テストグループ削除API\n・DELETE /api/test-groups/[groupId]",
            "viewpoint": "テストグループ削除時に関連データ（テストケース、テスト内容、テスト結果、ファイル等）が全てカスケード削除されること",
            "precondition": "・テストケース5件、テスト内容15件、テスト結果10件、エビデンス5件のテストグループが存在すること\n・システム管理者アカウントでログイン済みであること",
            "steps": [
                "1.テストグループ一覧から対象グループを選択し、削除を実行する",
                "2.テストグループ一覧で削除されたグループが表示されないことを確認する",
                "3.DBで関連テーブルのレコードが削除されていることを確認する",
                "4.S3上の関連ファイルが削除されていることを確認する",
            ],
            "expected": [
                "・削除確認ダイアログが表示され、削除が正常に完了すること",
                "・削除されたテストグループが一覧に表示されないこと",
                "・tt_test_cases, tt_test_contents, tt_test_results, tt_test_results_history, tt_test_evidences, tt_test_case_files, tt_test_group_tagsの関連レコードが全て削除されていること",
                "・S3上の関連ディレクトリ内のファイルが削除されていること",
            ],
        },
        # === E2Eシナリオ: セッション管理 ===
        {
            "major": "セッション管理シナリオ",
            "medium": "セッションタイムアウト",
            "minor": "",
            "type": "正常系",
            "spec": "・NextAuth.jsセッション管理\n・JWT有効期限",
            "viewpoint": "セッションタイムアウト後にログイン画面にリダイレクトされること",
            "precondition": "・一般ユーザアカウントでログイン済みであること\n・JWTトークンの有効期限が設定されていること",
            "steps": [
                "1.ログイン後、セッションタイムアウト時間まで操作せずに待機する",
                "2.タイムアウト後にテストグループ一覧画面にアクセスする",
            ],
            "expected": [
                "・セッションタイムアウト時間が経過すること",
                "・ログイン画面にリダイレクトされること\n・再ログイン後に正常に操作できること",
            ],
        },
    ]


def create_test_document(screen_id, doc_name, target_name, items, test_type, filename):
    """テスト試験書Excelファイルを作成"""
    wb = openpyxl.Workbook()

    # 表紙
    create_cover_sheet(wb, doc_name)

    # 改版履歴
    create_revision_sheet(wb)

    # 画面試験項目シート
    ws = wb.create_sheet("画面試験項目")
    setup_test_sheet(ws, screen_id, doc_name, target_name)
    write_test_items(ws, items, screen_id, test_type)

    # 保存
    output_path = os.path.join(os.path.dirname(__file__), filename)
    wb.save(output_path)
    print(f"Generated: {output_path}")
    return output_path


def main():
    docs = [
        {
            "screen_id": "ST01",
            "doc_name": "IT2_総合試験項目書_性能テスト",
            "target_name": "システム全体（性能テスト）",
            "items": get_performance_test_items(),
            "test_type": "IT2-PT",
            "filename": "IT2_総合試験項目書_性能テスト.xlsx",
        },
        {
            "screen_id": "ST02",
            "doc_name": "IT2_総合試験項目書_負荷テスト",
            "target_name": "システム全体（負荷テスト）",
            "items": get_load_test_items(),
            "test_type": "IT2-LT",
            "filename": "IT2_総合試験項目書_負荷テスト.xlsx",
        },
        {
            "screen_id": "ST03",
            "doc_name": "IT2_総合試験項目書_シナリオテスト",
            "target_name": "システム全体（シナリオテスト）",
            "items": get_scenario_test_items(),
            "test_type": "IT2-SC",
            "filename": "IT2_総合試験項目書_シナリオテスト.xlsx",
        },
    ]

    for doc in docs:
        create_test_document(**doc)

    print("\nAll documents generated successfully!")


if __name__ == "__main__":
    main()
